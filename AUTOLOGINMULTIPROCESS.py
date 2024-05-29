import tkinter as tk #VENTANAS EMERGENTES
import keyboard #MANDAR ACCION DE TECLAS
import time #RETARDOS, FUNCION DE TIME
import os #INTERACCION DE SISTEMAIMP
import unittest #TEST DE PRUEBAS UNITARIAS
import openpyxl
import multiprocessing
#import actions
#FROMS PARA USAR FUNCIONES
#COMMON .PYTHON
import gspread
from common import*
from openpyxl.styles import PatternFill # FORMATO DE CELDAS
from selenium.webdriver.support import expected_conditions as EC #EXPLICIT WAITS
from selenium.common.exceptions import NoSuchElementException #EXCEPTION para falla de encontrar elementos
from selenium.webdriver.support.ui import WebDriverWait #EXPLICIT WAITS
from selenium.common.exceptions import TimeoutException #USO DE TRY EXCEPT
from selenium.webdriver.chrome.options import Options #OPCIONES EXPERIMENTALES Y DE DESARROLLADOR
from selenium.webdriver.common.keys import Keys #USO DE ENVIO DE TECLAS
from selenium.webdriver.common.by import By #FUNCION BY PARA BUSQUEDA
from tkinter import simpledialog #DIALOGOS EN VENTANA EMERGENTE
from selenium import webdriver #IMPORTAR EL DRIVER DEL NAVEGADOR
#FIN DE IMPORTS
#DECLARACION DE VARIABLES DE TIEMPO
PAUSA = (10)
TIME = (1)
RETARD = (3)
URL = []
# Abrir IP en archivo Externo
ip = abrir_archivo_ip()
print("IP FOUNDED: ", ip)

class TestLogin(unittest.TestCase):

    def setUp(self):
        # Inicializar una lista de navegadores
        self.drivers = [webdriver.Chrome() for _ in range(10)]
        for driver in self.drivers:
            driver.implicitly_wait(10)  # Espera implícita de 10 segundos

    def tearDown(self):
        # Cerrar todos los navegadores al final de las pruebas
        for driver in self.drivers:
            driver.quit()

    def iniciar_sesion(self, driver, correo, contrasena):
        # Código para iniciar sesión con correo y contraseña en tu sitio web
        driver.get(ip)
        switch_to_frame(driver, "/html/frameset/frame", "FRAME 1")
        xpath_email(driver, "//*[@id='ContentPlaceHolder1_txtCorreoElectronico']", correo)
        xpath_password(driver, "//*[@id='ContentPlaceHolder1_txtContrasena']", contrasena)

    def iniciar_prueba(self, index, correos, contrasenas, resultados):
        for correo, contrasena in zip(correos, contrasenas):
            self.iniciar_sesion(self.drivers[index], correo, contrasena)

            try:
                WebDriverWait(self.drivers[index], 0.3).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/div/div/form/div[5]/iframe")))
                resultados[index].append("EXITOSO")
            except Exception as e:
                print(e)
                resultados[index].append("FALLIDO")

    def test_inicio_sesion_exitoso(self):
        # Cargar los correos y contraseñas desde el archivo Excel
        correos = []
        contrasenas = []

        archivo_excel = "LOG15.xlsx"
        workbook = openpyxl.load_workbook(archivo_excel)
        hoja = workbook.active

        for fila in hoja.iter_rows(min_row=2, values_only=True):
            if fila[0] is None:
                break
            correos.append(fila[1])
            contrasenas.append(fila[2])
        workbook.close()

        # Dividir los datos en 10 grupos
        chunk_size = len(correos) // 10
        data_chunks = [(correos[i:i + chunk_size], contrasenas[i:i + chunk_size]) for i in range(0, len(correos), chunk_size)]

        # Inicializar subarreglos para resultados
        resultados = [[] for _ in range(10)]

        # Ejecutar las pruebas en paralelo utilizando concurrent.futures
        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            futures = []
            for i in range(10):
                future = executor.submit(self.iniciar_prueba, i, data_chunks[i][0], data_chunks[i][1], resultados)
                futures.append(future)

            # Esperar a que todas las pruebas finalicen
            for future in futures:
                future.result()

        # Combinar los resultados en un arreglo final
        FINALRESULTS = []
        for resultado in resultados:
            FINALRESULTS.extend(resultado)

        print("Resultados finales:")
        print(FINALRESULTS)

if __name__ == "__main__":
    unittest.main()