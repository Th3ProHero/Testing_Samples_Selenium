#TESTING LOT OF LOGINS ELEMENTS
#PYTHON VERSION : 3.11.2
#SELENIUM VERSION : 4.8.3
#NAVEGADOR : CHROME V 111.0.5563.65 (Build oficial) (64 bits)
#SCRIPT BY DARK MAU V 3.1
#VELOCIDAD: Elementos evaluados: 589  ;  Tiempo transcurrido: 1831.993290424347 segundos
#IMPORTAR LIBRERIAS
import concurrent.futures #CONCURRENCIA Y PARALELISMO
import tkinter as tk #VENTANAS EMERGENTES
import keyboard #MANDAR ACCION DE TECLAS
import time #RETARDOS, FUNCION DE TIME
import os #INTERACCION DE SISTEMAIMP
import unittest #TEST DE PRUEBAS UNITARIAS
import openpyxl
import multiprocessing
#import actions
#FROMS PARA USAR FUNCIONES
#COMMON .PYTHON
import gspread
from common import*
from openpyxl.styles import PatternFill # FORMATO DE CELDAS
from selenium.webdriver.support import expected_conditions as EC #EXPLICIT WAITS
from selenium.common.exceptions import NoSuchElementException #EXCEPTION para falla de encontrar elementos
from selenium.webdriver.support.ui import WebDriverWait #EXPLICIT WAITS
from selenium.common.exceptions import TimeoutException #USO DE TRY EXCEPT
from selenium.webdriver.chrome.options import Options #OPCIONES EXPERIMENTALES Y DE DESARROLLADOR
from selenium.webdriver.common.keys import Keys #USO DE ENVIO DE TECLAS
from selenium.webdriver.common.by import By #FUNCION BY PARA BUSQUEDA
from tkinter import simpledialog #DIALOGOS EN VENTANA EMERGENTE
from selenium import webdriver #IMPORTAR EL DRIVER DEL NAVEGADOR
#FIN DE IMPORTS
#DECLARACION DE VARIABLES DE TIEMPO
PAUSA = (10)
TIME = (1)
RETARD = (3)
URL = []
# Abrir IP en archivo Externo
ip = abrir_archivo_ip()
print("IP FOUNDED: ", ip)

class PruebasInicioSesion(unittest.TestCase):
    def setUp(self):
        # Inicializar el navegador (ajusta esto según tu configuración)
        self.driver = webdriver.Chrome()
        self.driver.implicitly_wait(10)  # Espera implícita de 10 segundos

    def tearDown(self):
        # Cerrar el navegador al final de cada prueba
        self.driver.quit()

    def iniciar_sesion(self, correo, contrasena):
        # Código para iniciar sesión con correo y contraseña en tu sitio web
        driver = self.driver
        self.driver.get(ip)
        switch_to_frame(driver, "/html/frameset/frame", "FRAME 1")
        xpath_email(driver, "//*[@id='ContentPlaceHolder1_txtCorreoElectronico']", correo)
        xpath_password(driver, "//*[@id='ContentPlaceHolder1_txtContrasena']", contrasena)

    def test_inicio_sesion_exitoso(self):
        elementos_evaluados = 0
        tiempo_inicio = time.time()
        # Cargar los correos y contraseñas desde el archivo Excel
        correos = []
        contrasenas = []

        archivo_excel = "LOG11.xlsx"
        workbook = openpyxl.load_workbook(archivo_excel)
        hoja = workbook.active

        for fila in hoja.iter_rows(min_row=2, values_only=True):
            if fila[0] is None:
                break
            correos.append(fila[1])
            contrasenas.append(fila[2])
        print(correos)
        print(contrasenas)
        workbook.close()

        # Inicializar un arreglo para almacenar los resultados
        resultados = []

        # Iterar a través de los correos y contraseñas y realizar pruebas de inicio de sesión
        for correo, contrasena in zip(correos, contrasenas):
            self.iniciar_sesion(correo, contrasena)
            
            try:
                WebDriverWait(self.driver, 0.3).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/div/div/form/div[5]/iframe")))
                resultados.append("EXITOSO")
            except Exception as e:
                print(e)
                resultados.append("FALLIDO")
                
            elementos_evaluados += 1

        tiempo_transcurrido = time.time() - tiempo_inicio
        print(f"Elementos evaluados: {elementos_evaluados}")
        print(f"Tiempo transcurrido: {tiempo_transcurrido} segundos")
        # Abrir el archivo Excel nuevamente para escribir los resultados
        workbook = openpyxl.load_workbook(archivo_excel)
        hoja = workbook.active

        # Escribir los resultados en la columna D a partir de la segunda fila (D2 en adelante)
        for i, resultado in enumerate(resultados):
            hoja.cell(row=i+2, column=4, value=resultado)

        # Guardar el archivo Excel con los resultados
        workbook.save(archivo_excel)
        workbook.close()

if __name__ == "__main__":
    unittest.main()